"""RW-1: analytics completeness.

Covers the four additions: the point-biserial discrimination index, the live
during-assessment view, XLSX export, and Caliper-style event ingestion.
"""

from __future__ import annotations

import pytest
from events.caliper import EventAction, make_event
from pydantic import ValidationError

from app.domains.analytics.exports import to_xlsx
from app.domains.analytics.ingest import ingest_caliper, recent_events
from app.domains.analytics.service import (
    _point_biserial,
    growth_table,
    item_analysis,
    live_assessment,
    standards_table,
)
from app.domains.assessment.models import Assessment
from app.domains.attempts.models import Attempt
from app.domains.identity.models import User
from tests.conftest import AUTH


def test_point_biserial_sign_and_undefined_cases():
    # Strong students (high total) get it right, weak students wrong -> positive.
    positive = _point_biserial([(10.0, True), (9.0, True), (2.0, False), (1.0, False)])
    assert positive is not None and positive > 0

    # Reversed: strong students get it wrong -> negative (a miskeyed-item signal),
    # and it is NOT clamped to zero.
    negative = _point_biserial([(1.0, True), (2.0, True), (9.0, False), (10.0, False)])
    assert negative is not None and negative < 0

    # Undefined: everyone correct (no correctness variance), and too few points.
    assert _point_biserial([(5.0, True), (6.0, True)]) is None
    assert _point_biserial([(5.0, True)]) is None
    # No variance in totals -> undefined.
    assert _point_biserial([(5.0, True), (5.0, False)]) is None


async def test_item_analysis_carries_discrimination(db_session):
    rows = await item_analysis(db_session)
    assert rows, "expected seeded items"
    assert all("discrimination" in row for row in rows)


def test_xlsx_export_is_a_valid_workbook():
    data = to_xlsx("Item Analysis", ["A", "B"], [["x", 1], ["y", 2.5]])
    # XLSX is a zip container, so the bytes start with the PK signature.
    assert data[:2] == b"PK"
    import io

    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(data))
    sheet = book.active
    assert [c.value for c in sheet[1]] == ["A", "B"]
    # Numbers survive as numbers, not strings.
    assert sheet.cell(row=2, column=2).value == 1


def test_standards_and_growth_flatteners():
    s_headers, s_rows = standards_table(
        {"nodes": [{"code": "N1", "title": "T", "n_learners": 3, "avg_p_known": 0.5}]}
    )
    assert s_headers[0] == "Code" and s_rows[0][0] == "N1"
    g_headers, g_rows = growth_table(
        {"events": [{"t": "now", "node_id": "n", "correct": True, "p_known_after": 0.7}]}
    )
    assert g_headers[0] == "Time" and g_rows[0][2] == "yes"


async def test_live_assessment_counts_active_and_total(db_session):
    teacher = User(eureka_user_id="lt1", email="lt1@x.com", display_name="LT One")
    db_session.add(teacher)
    await db_session.flush()
    assessment = Assessment(title="Quiz", kind="quiz", created_by=teacher.id)
    db_session.add(assessment)
    await db_session.flush()
    db_session.add_all(
        [
            Attempt(user_id=teacher.id, assessment_id=assessment.id, status="in_progress",
                    answered_count=2, correct_count=1),
            Attempt(user_id=teacher.id, assessment_id=assessment.id, status="submitted",
                    answered_count=5, correct_count=5),
        ]
    )
    await db_session.flush()

    snapshot = await live_assessment(db_session, assessment.id)
    assert snapshot["n_total"] == 2
    assert snapshot["n_active"] == 1
    assert len(snapshot["attempts"]) == 2


async def test_caliper_ingestion_writes_and_reads_back(db_session):
    event = make_event("user-xyz", EventAction.submitted, "response", "resp-1", score=1.0)
    written = await ingest_caliper(db_session, [event.model_dump(mode="json")])
    assert written == 1

    recent = await recent_events(db_session, actor="user-xyz")
    assert len(recent) == 1
    assert recent[0]["action"] == "Submitted"
    assert recent[0]["extensions"].get("score") == 1.0


async def test_caliper_ingestion_rejects_malformed(db_session):
    with pytest.raises(ValidationError):
        await ingest_caliper(db_session, [{"actor": "no-action-or-object"}])


async def test_events_endpoint_ingest_open_query_teacher_only(client):
    event = make_event("user-abc", EventAction.viewed, "lesson", "lesson-1")
    resp = await client.post(
        "/api/v1/analytics/events", json=event.model_dump(mode="json"), headers=AUTH
    )
    assert resp.status_code == 200
    assert resp.json()["ingested"] == 1

    # Listing events is teacher-scoped; the mock identity is a student.
    forbidden = await client.get("/api/v1/analytics/events", headers=AUTH)
    assert forbidden.status_code == 403
